import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from app.services.pdf_service import T, _fmt, _offer_number, _format_date, _clean_offer_topic

# ─── Column layout (A–H, 8 columns) ──────────────────────────────────────────
# A=Nr(6)  B=Description(46)  C=spacer(3)  D=spacer(3)  E=spacer(3)
# F=Buc(7)  G=UnitPrice(14)  H=Total(14)
# Description area B:E merged = 55 chars   Logo anchored at D in company header

_COL_A, _COL_B, _COL_E, _COL_F, _COL_G, _COL_H = 1, 2, 5, 6, 7, 8
_GREY   = "D9D9D9"
_LIGHT  = "F5F5F5"


def _font(bold=False, size=9, color="000000", name="Arial",
          underline=None, italic=False):
    return Font(name=name, size=size, bold=bold, color=color,
                underline=underline, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _box():
    t = Side(style="thin", color="000000")
    return Border(left=t, right=t, top=t, bottom=t)


def _box_name_row():
    """Top/left/right borders — no bottom, so the description rows below show no separator."""
    b = Side(style="thin", color="000000")
    return Border(left=b, right=b, top=b)


def _box_desc_mid():
    """Left/right borders only — no top/bottom, removes internal lines between description rows."""
    b = Side(style="thin", color="000000")
    return Border(left=b, right=b)


def _box_desc_last():
    """Left/right/bottom borders — no top, closes the product block at the bottom."""
    b = Side(style="thin", color="000000")
    return Border(left=b, right=b, bottom=b)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set(ws, row, col, value="", font=None, fill=None,
         align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font          = font
    if fill:    c.fill          = fill
    if align:   c.alignment     = align
    if border:  c.border        = border
    if num_fmt: c.number_format = num_fmt
    return c


def _merge(ws, r1, c1, r2, c2):
    """Merge cells and return the top-left cell."""
    ws.merge_cells(
        start_row=r1, start_column=c1,
        end_row=r2,   end_column=c2,
    )
    return ws.cell(row=r1, column=c1)


def _uri_stream(data_uri: str) -> io.BytesIO | None:
    try:
        _, data = data_uri.split(",", 1)
        return io.BytesIO(base64.b64decode(data))
    except Exception:
        return None


def _xl_image(stream: io.BytesIO, max_w: int, max_h: int) -> XLImage | None:
    try:
        stream.seek(0)
        pil = PILImage.open(stream)
        scale = min(max_w / pil.width, max_h / pil.height, 1.0)
        img = XLImage(stream)
        img.width  = int(pil.width  * scale)
        img.height = int(pil.height * scale)
        return img
    except Exception:
        return None


# ─── Main generator ───────────────────────────────────────────────────────────

def generate_offer_excel(
    quote: dict,
    company: dict | None,
    products: list[dict],
    lang: str,
    logo_data_uri: str | None = None,
    signature_data_uri: str | None = None,
) -> bytes:
    tr           = T[lang]
    today        = datetime.now()
    offer_no     = _offer_number(quote["id"], today)
    date_str     = _format_date(today, lang)
    client_name  = company["name"] if company else ""
    items        = quote.get("items") or []
    installation = float(quote.get("installation") or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Offer"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    row = 1

    # ── Company header – plain white, text in A, logo in D area ───────────────
    ws.row_dimensions[row].height = 18  # row 1, ≥16
    _set(ws, row, _COL_A, "S.C. SMSS REITLER S.R.L.",
         font=_font(bold=False, size=10, name="Times New Roman"), align=_align())
    logo_start = row
    row += 1

    for line in [
        "445100 Carei, jud. Satu Mare – RO, Calea Armatei Române, nr. 90",
        "Tel./Fax.: 00-40-261-863430, Mobil: 0744-520219",
        "CIF: RO15478578",
        "Nr. Reg. Com.: J30/427/2003",
        "Cont LEI: RO33BTRL03101202N12327XX",
        "Cont EUR: RO08BTRL03104202N12327XX",
        "Banca: Banca Transilvania, Agenția Carei",
        "E-mail: smsreitler@gmail.com",
        "Web: www.smsreitler.ro",
    ]:
        ws.row_dimensions[row].height = 16  # rows 2–10, ≥16 for logo height
        _set(ws, row, _COL_A, line, font=_font(size=8.5), align=_align())
        row += 1

    # Embed logo anchored at column C, top row of company section
    if logo_data_uri:
        stream = _uri_stream(logo_data_uri)
        if stream:
            img = _xl_image(stream, max_w=360, max_h=300)
            if img:
                ws.add_image(img, f"C{logo_start}")

    row += 1  # empty spacer (row 11)

    # ── Rows 12-13: Client / Date / Reg — client merged vertically ───────────
    ws.row_dimensions[row].height     = 16
    ws.row_dimensions[row + 1].height = 14

    c = _merge(ws, row, _COL_A, row + 1, _COL_F)   # A12:F13 merged
    c.value     = f"{tr['to']} {client_name},"
    c.font      = _font(bold=True, size=12)
    c.alignment = _align("center", v="center")
    c.border    = _box()

    _set(ws, row,     _COL_G, tr["offerDate"],       # G12
         font=_font(bold=True,  size=11), align=_align("right"), border=_box())
    _set(ws, row,     _COL_H, date_str,              # H12
         font=_font(bold=False, size=11), align=_align(),         border=_box())
    _set(ws, row + 1, _COL_G, tr["regNo"],           # G13
         font=_font(bold=True,  size=11), align=_align("right"), border=_box())
    _set(ws, row + 1, _COL_H, offer_no,              # H13
         font=_font(bold=False, size=11), align=_align(),         border=_box())
    row += 2

    # ── Row 14: Offer title ────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    c = _merge(ws, row, _COL_A, row, _COL_H)
    offer_topic = _clean_offer_topic(quote.get('name', ''), client_name, tr["offerFor"])
    c.value     = f"{tr['offerFor']} {offer_topic}".strip()
    c.font      = _font(bold=True, size=13)
    c.alignment = _align("center")
    c.border    = _box()
    row += 1

    # ── Intro ──────────────────────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 14
    c = _merge(ws, row, _COL_A, row, _COL_H)
    c.value     = tr["intro"]
    c.font      = _font(size=12)
    c.alignment = _align(wrap=True)
    row += 1

    # (no spacer before table)

    # ── Table header – 2 rows ──────────────────────────────────────────────────
    th1, th2 = row, row + 1
    for r in (th1, th2):
        ws.row_dimensions[r].height = 15

    hf = _font(bold=True, size=9)
    hfill = _fill(_GREY)

    # A: "Nr." / "Crt." – two separate cells (not merged)
    _set(ws, th1, _COL_A, "Nr.",   font=hf, fill=hfill, align=_align("center"), border=_box())
    _set(ws, th2, _COL_A, "Crt.",  font=hf, fill=hfill, align=_align("center"), border=_box())

    # B:E "Denumire reper" – spans both rows
    c = _merge(ws, th1, _COL_B, th2, 5)
    c.value     = tr["colName"]
    c.font      = hf
    c.fill      = hfill
    c.alignment = _align("center")
    c.border    = _box()

    # F "Buc" – spans both rows
    c = _merge(ws, th1, _COL_F, th2, _COL_F)
    c.value     = tr["colQty"]
    c.font      = hf
    c.fill      = hfill
    c.alignment = _align("center")
    c.border    = _box()

    # G "Preț unitar" – spans both rows
    c = _merge(ws, th1, _COL_G, th2, _COL_G)
    c.value     = tr["colUnit"].replace("\n", " ")
    c.font      = hf
    c.fill      = hfill
    c.alignment = _align("center", wrap=True)
    c.border    = _box()

    # H "Preț total" – spans both rows
    c = _merge(ws, th1, _COL_H, th2, _COL_H)
    c.value     = tr["colTotal"].replace("\n", " ")
    c.font      = hf
    c.fill      = hfill
    c.alignment = _align("center", wrap=True)
    c.border    = _box()

    row = th2 + 1

    # ── Item rows ──────────────────────────────────────────────────────────────
    grand_total = 0.0

    for idx, item in enumerate(items):
        product    = next((p for p in products if p["id"] == item.get("productId")), None)
        prod_name  = product["name"] if product else item.get("productId", "")
        desc_map   = (product.get("description") or {}) if product else {}
        lang_key   = lang if lang in desc_map else "ro"
        desc_raw   = desc_map.get(lang_key, "")
        notes      = (item.get("notes") or "").strip()
        qty        = item.get("quantity", 0)
        unit_price = float(item.get("unitPrice", 0))
        line_total = qty * unit_price
        grand_total += line_total

        # Split description into individual lines
        desc_lines = [l.strip() for l in desc_raw.split("\n") if l.strip()]
        if notes:
            desc_lines.append(notes)

        # Total rows this item occupies: 1 (name) + N (desc lines)
        n = 1 + len(desc_lines)
        r0 = row  # first row of this item

        # Merge A, F, G, H vertically across all n rows
        if n > 1:
            _merge(ws, r0, _COL_A, r0 + n - 1, _COL_A)
            _merge(ws, r0, _COL_F, r0 + n - 1, _COL_F)
            _merge(ws, r0, _COL_G, r0 + n - 1, _COL_G)
            _merge(ws, r0, _COL_H, r0 + n - 1, _COL_H)

        # Nr. cell
        ws.row_dimensions[r0].height = 15
        _set(ws, r0, _COL_A, idx + 1,
             font=_font(size=9), align=_align("center"), border=_box())

        # Product name row (bold) — open bottom if description lines follow
        c = _merge(ws, r0, _COL_B, r0, 5)
        c.value     = prod_name
        c.font      = _font(bold=True, size=12)
        c.alignment = _align()
        c.border    = _box() if not desc_lines else _box_name_row()

        # Description lines – one per row, no internal borders
        for i, dl in enumerate(desc_lines):
            r = r0 + 1 + i
            is_last = (i == len(desc_lines) - 1)
            ws.row_dimensions[r].height = 19 if is_last else 13
            c = _merge(ws, r, _COL_B, r, 5)
            c.value     = dl
            c.font      = _font(size=12)
            c.alignment = _align()
            c.border    = _box_desc_last() if is_last else _box_desc_mid()



        # Qty / prices (merged vertically, values on first row)
        _set(ws, r0, _COL_F, qty,
             font=_font(size=12), align=_align("center"), border=_box())
        _set(ws, r0, _COL_G, unit_price,
             font=_font(size=12), align=_align("center"), border=_box(), num_fmt='#,##0.00')
        _set(ws, r0, _COL_H, line_total,
             font=_font(size=12), align=_align("center"), border=_box(), num_fmt='#,##0.00')

        row += n

    # ── Installation row ───────────────────────────────────────────────────────
    if installation > 0:
        grand_total += installation
        r0 = row
        n = 2  # name row + sub-label row

        _merge(ws, r0, _COL_A, r0 + n - 1, _COL_A)
        _merge(ws, r0, _COL_F, r0 + n - 1, _COL_F)
        _merge(ws, r0, _COL_G, r0 + n - 1, _COL_G)
        _merge(ws, r0, _COL_H, r0 + n - 1, _COL_H)

        ws.row_dimensions[r0].height = 15
        _set(ws, r0, _COL_A, len(items) + 1,
             font=_font(size=9), align=_align("center"), border=_box())

        c = _merge(ws, r0, _COL_B, r0, 5)
        c.value     = tr["installation"]
        c.font      = _font(bold=True, size=9)
        c.alignment = _align()
        c.border    = _box_name_row()

        ws.row_dimensions[r0 + 1].height = 19
        c = _merge(ws, r0 + 1, _COL_B, r0 + 1, 5)
        c.value     = tr["installationSub"]
        c.font      = _font(size=8.5)
        c.alignment = _align()
        c.border    = _box_desc_last()

        _set(ws, r0, _COL_F, 1,
             font=_font(size=9), align=_align("center"), border=_box())
        _set(ws, r0, _COL_G, installation,
             font=_font(size=9), align=_align("center"), border=_box(), num_fmt='#,##0.00')
        _set(ws, r0, _COL_H, installation,
             font=_font(size=9), align=_align("center"), border=_box(), num_fmt='#,##0.00')

        row += n

    # ── Grand total ────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    # Label merge A:E (se termină înainte de Buc)
    c = _merge(ws, row, _COL_A, row, _COL_E)
    c.value     = tr["grandTotal"]
    c.font      = _font(bold=True, size=13)
    c.fill      = _fill(_LIGHT)
    c.alignment = _align("right")
    c.border    = _box()

    # Valoarea merge F:H, centrat
    c2 = _merge(ws, row, _COL_F, row, _COL_H)
    c2.value      = grand_total
    c2.font       = _font(bold=True, size=13)
    c2.fill       = _fill(_LIGHT)
    c2.alignment  = _align("center")
    c2.border     = _box()
    c2.number_format = '#,##0.00'
    row += 2

    # ── Price note ─────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    c = _merge(ws, row, _COL_A, row, _COL_H)
    c.value     = tr["priceNote"]
    c.font      = _font(bold=True, size=13)
    c.alignment = _align(wrap=True)
    row += 2

    # ── Conditions title ───────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    c = _merge(ws, row, _COL_A, row, _COL_H)
    c.value     = tr["conditionsTitle"]
    c.font      = _font(bold=True, size=12, underline="single")
    c.alignment = _align("center")
    row += 1

    row += 1  # spacer

    # Helper: full-width merged line
    def _full(text, bold=True, h=14, italic=False, size=12):
        nonlocal row
        ws.row_dimensions[row].height = h
        c = _merge(ws, row, _COL_A, row, _COL_H)
        c.value     = text
        c.font      = _font(bold=bold, italic=italic, size=size)
        c.alignment = _align(wrap=True)
        row += 1

    _full(f"{tr['deliveryLabel']}  {quote.get('deliveryTimeWeeks', 4)} {tr['deliveryWeeks']}",
          bold=True)
    row += 1

    # ── Not-included list ──────────────────────────────────────────────────────
    _set(ws, row, _COL_A, tr["notIncludeLabel"],
         font=_font(bold=True, size=12), align=_align())
    row += 1
    for excl in tr["notInclude"]:
        ws.row_dimensions[row].height = 13
        c = _merge(ws, row, _COL_A, row, _COL_H)
        c.value     = f"- {excl}"
        c.font      = _font(size=10)
        c.alignment = _align(wrap=True)
        row += 1

    row += 1

    # ── Warranty ───────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    c = _merge(ws, row, _COL_A, row, _COL_H)
    c.value     = tr["warrantyLabel"]
    c.font      = _font(bold=True, size=11)
    c.alignment = _align()
    row += 1

    ws.row_dimensions[row].height = 14
    c = _merge(ws, row, _COL_A, row, _COL_H)
    c.value     = tr["warrantyText"]
    c.font      = _font(size=11)
    c.alignment = _align()
    row += 1

    row += 2  # blank rows before service text
    _full(tr["serviceText"], bold=False, size=11)
    row += 2  # blank rows after service text

    # ── Payment ────────────────────────────────────────────────────────────────
    _full(tr["paymentLabel"], bold=True, size=12)
    for pay in [tr["payment1"], tr["payment2"], tr["payment3"]]:
        ws.row_dimensions[row].height = 13
        c = _merge(ws, row, _COL_B, row, _COL_H)
        c.value     = pay
        c.font      = _font(size=11)
        c.alignment = _align()
        row += 1

    row += 1

    # ── Validity ───────────────────────────────────────────────────────────────
    _full(f"{tr['validityLabel']}  {tr['validity21']}", bold=True, size=12)
    row += 1

    # ── Closing ────────────────────────────────────────────────────────────────
    for line in [tr["before"], tr["contact"], tr["review"]]:
        _full(line, bold=False, h=13, size=11)

    row += 1

    # ── Signature ──────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 13
    _set(ws, row, _COL_A, tr["regards"], font=_font(size=11), align=_align())
    row += 1

    ws.row_dimensions[row].height = 14
    _set(ws, row, _COL_A, "Alexandru Reitler", font=_font(bold=True, size=11), align=_align())
    row += 1

    if signature_data_uri:
        stream = _uri_stream(signature_data_uri)
        if stream:
            sig_img = _xl_image(stream, max_w=240, max_h=180)
            if sig_img:
                ws.add_image(sig_img, f"B{row}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
